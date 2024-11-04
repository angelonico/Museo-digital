import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Rutas
path_carpeta = r'C:\Users\victo\OneDrive\Escritorio\Colecciones Desk\Scripts'
path_destino = path_carpeta + r"\Reporte\Errores.xlsx"

def Reportes_F2(df):
  
  # Guardar el DataFrame a un archivo Excel
  df.to_excel(path_destino, index=False)

  # Cargar el archivo Excel con openpyxl
  wb = load_workbook(path_destino)
  ws = wb.active

  # Definir el color de fondo (rojo)
  fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

  # Aplicar el color a una fila específica, por ejemplo la fila 2
  row_number = 2
  for cell in ws[row_number]:
      cell.fill = fill

  # Guardar el archivo con formato aplicado
  wb.save(path_destino)
